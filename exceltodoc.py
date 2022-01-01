


# importing openpyxl module
import openpyxl
from docxtpl import DocxTemplate
from docx.shared import Pt
empty=''
# Give the location of the file
path = "myexcel.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
m_col = sheet_obj.max_column
# Loop will print all values
# of first column
mydict={}
myarray=[]
for j in range(1, m_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = j).value
    mydict[cell_obj]=""
    myarray.append(cell_obj)

print(mydict)
print(myarray)

for i in range(1, m_row + 1):
    array=[]
    for j in range(1, m_col + 1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        #print(cell_obj.value,end="|")
        array.append(cell_obj.value)
    doc = DocxTemplate("template1.docx")
    style = doc.styles['Normal']
    font = style.font    
    font.name = 'Calibri'
    font.italic = True
    font.size = Pt(12)
    context = {'ID': empty,
           'Start time': empty,
           'Completion time': empty,
           'Email': empty,
           'Name': empty,
           'Faculty': empty,
           'Department': empty,
           'Year': empty,
           'Semester': empty,
           'Module_Name': empty,
           'Module_Code': empty,
           'Credit_Hours': empty,
           'Prerequisite': empty,
           'Lecturer': empty,
           'Email2': empty,
           'Office_hours': empty,
           'Lecture_timetable': empty,
           'Required_Textbooks': empty,
           'References': empty,
           'Module_Short_Description': empty,
           'Module_Objectives': empty,
           'Week_by_Week_Course_Content': empty,
           'Knowledge_and_Understanding': empty,
           'Intellectual_Skills': empty,
           'Professional_Skills': empty,
           'General_Skills': empty,
           'Teaching_and_learning_Methods': empty,
           'Assessment_Methods': empty,
           'Module_development_plans': empty
           }
    for x in range(len(array)):
        key = list(context)[x]
        array[x]=str(array[x])
        print(myarray[x]," : ",array[x])
        context[key]=myarray[x]+" : \n"+array[x]+"\n"
    doc.render(context)
    doc.save("genrated_word/"+array[10]+".docx")
    print()
